from __future__ import annotations

import logging
from copy import copy
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from .utils import formatar_valor_brl


COLUNAS = [
    "Data",
    "Status",
    "Fornecedor",
    "CNPJ",
    "NF",
    "Pedido(s)",
    "Valor Pedido(s)",
    "Valor NF",
    "Valor Produtos NF",
    "Valor Frete",
    "Outras Despesas",
    "Valor Boleto(s)",
    "Diferenca",
    "Documentos Encontrados",
    "Pasta Origem",
    "Pasta Destino",
    "Vencimento(s)",
    "Tipo Documento",
    "Arquivo Original",
    "Arquivo Final",
    "Observacoes",
    "Erro Encontrado",
    "Confianca OCR",
    "Confianca Documento",
    "Confianca Processo",
    "Score Auditoria",
    "Risco Auditoria",
    "Critérios Score",
]

COLUNAS_PENDENCIAS = [
    "Data Processamento",
    "ID Processo",
    "Status",
    "Motivo Principal",
    "Motivos Detalhados",
    "Ação Recomendada",
    "Fornecedor",
    "CNPJ Fornecedor",
    "NF",
    "Pedido",
    "Valor Pedido",
    "Valor NF",
    "Valor Boleto",
    "Diferença Pedido x NF",
    "Diferença NF x Boleto",
    "Documentos Encontrados",
    "Documentos Faltantes",
    "Arquivos Envolvidos",
    "Pasta Pendência",
    "Origem OCR/Digital",
    "Confiança",
    "Score Auditoria",
    "Risco Auditoria",
    "Critérios Score",
    "Observação",
]

MOTIVOS_PRINCIPAIS = {
    "PENDENTE_NF": "Nota fiscal nao identificada",
    "PENDENTE_BOLETO": "Boleto nao identificado",
    "PENDENTE_PEDIDO": "Pedido de compra nao identificado",
    "PENDENTE_CNPJ": "Divergencia de CNPJ",
    "PENDENTE_VALOR": "Divergencia de valor",
    "PENDENTE_OCR": "OCR com baixa confianca",
    "PENDENTE_CLASSIFICACAO": "Documento nao classificado",
    "PENDENTE_ERRO_ARQUIVO": "Erro ao acessar arquivo fisico",
    "VALOR_EXCEDENTE": "Valor excede o pedido",
    "APROVADO_PARCIAL": "Processo aprovado parcialmente",
}

ACOES_RECOMENDADAS = {
    "PENDENTE_NF": "Anexar a nota fiscal correspondente ao pedido/processo.",
    "PENDENTE_BOLETO": "Anexar o boleto correspondente a nota fiscal.",
    "PENDENTE_PEDIDO": "Anexar o pedido de compra ou verificar se o fornecedor e excecao.",
    "PENDENTE_CNPJ": "Conferir se os documentos pertencem ao mesmo fornecedor/processo.",
    "PENDENTE_VALOR": "Conferir valores entre pedido, nota fiscal e boleto.",
    "PENDENTE_OCR": "Revisar manualmente o PDF escaneado ou substituir por arquivo digital.",
    "PENDENTE_CLASSIFICACAO": "Verificar o tipo do documento e renomear/adicionar manualmente se necessario.",
    "PENDENTE_ERRO_ARQUIVO": "Verificar se o arquivo existe fisicamente na pasta ENTRADA/SharePoint.",
    "VALOR_EXCEDENTE": "Conferir se ha pedido complementar ou erro no valor da NF.",
    "APROVADO_PARCIAL": "Verificar documentos restantes antes do envio fiscal.",
}


def gerar_relatorios(processos: list[dict], pasta_relatorios: Path, data_ref) -> tuple[Path, Path, Path]:
    pasta_relatorios.mkdir(parents=True, exist_ok=True)
    linhas = montar_linhas(processos)
    linhas_pendencias = montar_linhas_pendencias(processos)
    sufixo = data_ref.strftime("%d-%m-%Y")
    caminho_xlsx = pasta_relatorios / f"relatorio_processamento_{sufixo}.xlsx"
    caminho_pendencias = pasta_relatorios / f"relatorio_pendencias_{sufixo}.xlsx"
    caminho_txt = pasta_relatorios / f"resumo_processamento_{sufixo}.txt"

    try:
        gerar_excel(linhas, caminho_xlsx)
    except PermissionError:
        caminho_xlsx = caminho_unico(caminho_xlsx)
        gerar_excel(linhas, caminho_xlsx)

    try:
        gerar_excel_pendencias(linhas_pendencias, caminho_pendencias)
    except PermissionError:
        caminho_pendencias = caminho_unico(caminho_pendencias)
        gerar_excel_pendencias(linhas_pendencias, caminho_pendencias)

    try:
        gerar_txt(processos, linhas, caminho_txt)
    except PermissionError:
        caminho_txt = caminho_unico(caminho_txt)
        gerar_txt(processos, linhas, caminho_txt)
    return caminho_xlsx, caminho_pendencias, caminho_txt


def montar_linhas(processos: list[dict]) -> list[dict]:
    linhas = []
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for processo in processos:
        docs = processo["documentos"]
        vencimentos = ", ".join(sorted({d["vencimento"] for d in docs if d.get("vencimento")}))
        pedidos = ", ".join(sorted({str(d["numero_pedido"]) for d in docs if d.get("numero_pedido")}))
        erro = "; ".join(processo.get("erros", []))
        tipos = ", ".join(sorted({str(d.get("tipo_documento")) for d in docs if d.get("tipo_documento")}))
        pasta_origem = primeiro_texto(Path(d["arquivo_origem"]).parent for d in docs if d.get("arquivo_origem"))
        pasta_destino = primeiro_texto(Path(str(d["arquivo_final"])).parent for d in docs if d.get("arquivo_final"))
        diferenca = calcular_diferenca(processo)
        for doc in docs:
            linhas.append(
                {
                    "Data": agora,
                    "Status": processo.get("status"),
                    "Fornecedor": doc.get("fornecedor_nome"),
                    "CNPJ": doc.get("fornecedor_cnpj"),
                    "NF": doc.get("numero_nf"),
                    "Pedido(s)": pedidos,
                    "Valor Pedido(s)": _fmt(processo.get("valor_pedidos")),
                    "Valor NF": _fmt(processo.get("valor_nf")),
                    "Valor Produtos NF": _fmt(processo.get("valor_produtos_nf")),
                    "Valor Frete": _fmt(doc.get("valor_frete")),
                    "Outras Despesas": _fmt(doc.get("outras_despesas")),
                    "Valor Boleto(s)": _fmt(processo.get("valor_boletos")),
                    "Diferenca": _fmt(diferenca),
                    "Documentos Encontrados": tipos,
                    "Pasta Origem": str(pasta_origem) if pasta_origem else None,
                    "Pasta Destino": str(pasta_destino) if pasta_destino else None,
                    "Vencimento(s)": vencimentos,
                    "Tipo Documento": doc.get("tipo_documento"),
                    "Arquivo Original": doc.get("arquivo_nome"),
                    "Arquivo Final": doc.get("arquivo_final"),
                    "Observacoes": doc.get("observacoes"),
                    "Erro Encontrado": erro,
                    "Confianca OCR": doc.get("confianca_extracao"),
                    "Confianca Documento": doc.get("confianca_documento"),
                    "Confianca Processo": processo.get("confianca_processo"),
                    "Score Auditoria": processo.get("score_auditoria"),
                    "Risco Auditoria": processo.get("risco_auditoria"),
                    "Critérios Score": criterios_score(processo),
                }
            )
    return linhas


def montar_linhas_pendencias(processos: list[dict]) -> list[dict]:
    linhas = []
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for processo in processos:
        if processo.get("status") == "APROVADO" and not processo.get("erros"):
            continue
        try:
            linhas.append(montar_linha_pendencia(processo, agora))
        except Exception as exc:
            logging.exception("Falha ao montar linha do relatorio de pendencias: %s", exc)
    return linhas


def montar_linha_pendencia(processo: dict, data_processamento: str) -> dict:
    docs = processo.get("documentos", [])
    status = processo.get("status") or "PENDENTE_CLASSIFICACAO"
    valor_pedido = processo.get("valor_pedidos")
    valor_nf = processo.get("valor_nf")
    valor_boleto = processo.get("valor_boletos")
    return {
        "Data Processamento": data_processamento,
        "ID Processo": processo.get("id"),
        "Status": status,
        "Motivo Principal": motivo_principal(status),
        "Motivos Detalhados": "; ".join(str(erro) for erro in processo.get("erros", []) if erro),
        "Ação Recomendada": acao_recomendada(status),
        "Fornecedor": primeiro_texto(doc.get("fornecedor_nome") for doc in docs),
        "CNPJ Fornecedor": primeiro_texto(doc.get("fornecedor_cnpj") for doc in docs),
        "NF": "; ".join(unicos(doc.get("numero_nf") for doc in docs)),
        "Pedido": "; ".join(unicos(doc.get("numero_pedido") for doc in docs)),
        "Valor Pedido": valor_pedido,
        "Valor NF": valor_nf,
        "Valor Boleto": valor_boleto,
        "Diferença Pedido x NF": diferenca_decimal(valor_pedido, valor_nf),
        "Diferença NF x Boleto": diferenca_decimal(valor_nf, valor_boleto),
        "Documentos Encontrados": "; ".join(tipos_documentos(docs)),
        "Documentos Faltantes": "; ".join(documentos_faltantes(status, docs)),
        "Arquivos Envolvidos": "; ".join(arquivos_envolvidos(docs)),
        "Pasta Pendência": pasta_pendencia(docs),
        "Origem OCR/Digital": "; ".join(unicos(doc.get("origem_texto") for doc in docs)),
        "Confiança": processo.get("confianca_processo") or primeiro_texto(doc.get("confianca_documento") for doc in docs),
        "Score Auditoria": processo.get("score_auditoria"),
        "Risco Auditoria": processo.get("risco_auditoria"),
        "Critérios Score": criterios_score(processo),
        "Observação": "; ".join(unicos(doc.get("observacoes") for doc in docs)),
    }


def primeiro_texto(valores) -> object | None:
    for valor in valores:
        if valor not in (None, ""):
            return valor
    return None


def calcular_diferenca(processo: dict) -> Decimal | None:
    nf = processo.get("valor_nf")
    pedido = processo.get("valor_pedidos")
    boleto = processo.get("valor_boletos")
    referencia = boleto if boleto is not None else pedido
    if nf is None or referencia is None:
        return None
    return Decimal(nf) - Decimal(referencia)


def diferenca_decimal(esquerda, direita) -> Decimal | None:
    if esquerda is None or direita is None:
        return None
    try:
        return Decimal(esquerda) - Decimal(direita)
    except Exception:
        logging.exception("Falha ao calcular diferenca do relatorio de pendencias.")
        return None


def motivo_principal(status: str) -> str:
    return MOTIVOS_PRINCIPAIS.get(status, "Pendencia nao classificada")


def acao_recomendada(status: str) -> str:
    return ACOES_RECOMENDADAS.get(status, "Revisar manualmente o processo e os documentos anexados.")


def unicos(valores) -> list[str]:
    vistos = set()
    resultado = []
    for valor in valores:
        if valor in (None, ""):
            continue
        texto = str(valor)
        if texto not in vistos:
            vistos.add(texto)
            resultado.append(texto)
    return resultado


def tipos_documentos(docs: list[dict]) -> list[str]:
    return sorted(unicos(doc.get("tipo_documento") for doc in docs))


def documentos_faltantes(status: str, docs: list[dict]) -> list[str]:
    encontrados = set(tipos_documentos(docs))
    faltantes = {
        "PENDENTE_NF": ["NF_PRODUTO/NF_SERVICO"],
        "PENDENTE_BOLETO": ["BOLETO"],
        "PENDENTE_PEDIDO": ["PEDIDO_COMPRA"],
        "PENDENTE_CLASSIFICACAO": ["DOCUMENTO_CLASSIFICADO"],
    }.get(status, [])
    return [documento for documento in faltantes if documento not in encontrados]


def arquivos_envolvidos(docs: list[dict]) -> list[str]:
    nomes = []
    for doc in docs:
        nome = doc.get("arquivo_nome")
        if not nome and doc.get("arquivo_origem"):
            nome = Path(str(doc.get("arquivo_origem"))).name
        if not nome and doc.get("caminho_original"):
            nome = Path(str(doc.get("caminho_original"))).name
        nomes.append(nome)
    return unicos(nomes)


def pasta_pendencia(docs: list[dict]) -> str | None:
    for doc in docs:
        arquivo_final = doc.get("arquivo_final")
        if arquivo_final:
            return str(Path(str(arquivo_final)).parent)
    return None


def criterios_score(processo: dict) -> str:
    return "; ".join(str(criterio) for criterio in processo.get("criterios_score", []) if criterio)


def gerar_excel(linhas: list[dict], caminho: Path) -> None:
    try:
        import pandas as pd

        pd.DataFrame(linhas, columns=COLUNAS).to_excel(caminho, index=False)
    except ImportError:
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Conferencia"
            ws.append(COLUNAS)
            for linha in linhas:
                ws.append([linha.get(coluna) for coluna in COLUNAS])
            wb.save(caminho)
        except ImportError:
            gerar_xlsx_minimo(linhas, caminho)


def gerar_excel_pendencias(linhas: list[dict], caminho: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        gerar_xlsx_minimo(linhas, caminho, COLUNAS_PENDENCIAS)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendencias"
    ws.append(COLUNAS_PENDENCIAS)
    for linha in linhas:
        valores = []
        for coluna in COLUNAS_PENDENCIAS:
            try:
                valores.append(linha.get(coluna))
            except Exception as exc:
                logging.exception("Falha ao preencher celula '%s' no relatorio de pendencias: %s", coluna, exc)
                valores.append(None)
        ws.append(valores)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    status_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    money_columns = {
        "Valor Pedido",
        "Valor NF",
        "Valor Boleto",
        "Diferença Pedido x NF",
        "Diferença NF x Boleto",
    }
    wrap_columns = {"Motivos Detalhados", "Ação Recomendada", "Arquivos Envolvidos", "Critérios Score", "Observação"}
    for column_index, coluna in enumerate(COLUNAS_PENDENCIAS, start=1):
        letra = get_column_letter(column_index)
        largura = min(max(len(coluna) + 2, 14), 45)
        if coluna in wrap_columns:
            largura = 45
        ws.column_dimensions[letra].width = largura
        for cell in ws[letra][1:]:
            if coluna in money_columns and cell.value is not None:
                cell.number_format = '"R$" #,##0.00'
            if coluna in wrap_columns:
                alinhamento = copy(cell.alignment)
                alinhamento.wrap_text = True
                cell.alignment = alinhamento
            if coluna == "Status" and cell.value:
                cell.fill = status_fill

    if ws.max_row > 1:
        tabela = Table(displayName="TabelaPendencias", ref=ws.dimensions)
        tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tabela)
    wb.save(caminho)


def gerar_txt(processos: list[dict], linhas: list[dict], caminho: Path) -> None:
    aprovados = sum(1 for p in processos if p.get("status") == "APROVADO")
    pendentes = len(processos) - aprovados
    with caminho.open("w", encoding="utf-8") as arquivo:
        arquivo.write("RELATORIO DE CONFERENCIA PDF\n")
        arquivo.write(f"Processos aprovados: {aprovados}\n")
        arquivo.write(f"Processos pendentes: {pendentes}\n\n")
        for processo in processos:
            arquivo.write(f"{processo['id']} - {processo.get('status')}\n")
            if processo.get("erros"):
                arquivo.write("Erros: " + "; ".join(processo["erros"]) + "\n")
            for doc in processo["documentos"]:
                arquivo.write(f"  - {doc.get('tipo_documento')} | {doc.get('arquivo_nome')} | pagina {doc.get('pagina')}\n")
            arquivo.write("\n")


def _fmt(valor: Decimal | None) -> str | None:
    return formatar_valor_brl(valor) if valor is not None else None


def gerar_xlsx_minimo(linhas: list[dict], caminho: Path, colunas: list[str] | None = None) -> None:
    colunas = colunas or COLUNAS
    rows = [colunas] + [[linha.get(coluna) for coluna in colunas] for linha in linhas]
    sheet_rows = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for col_index, value in enumerate(row, start=1):
            ref = f"{_coluna_excel(col_index)}{row_index}"
            texto = escape("" if value is None else str(value))
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{texto}</t></is></c>')
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')

    worksheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )
    workbook = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Conferencia" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    with ZipFile(caminho, "w", ZIP_DEFLATED) as xlsx:
        xlsx.writestr("[Content_Types].xml", content_types)
        xlsx.writestr("_rels/.rels", rels)
        xlsx.writestr("xl/workbook.xml", workbook)
        xlsx.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        xlsx.writestr("xl/worksheets/sheet1.xml", worksheet)


def _coluna_excel(indice: int) -> str:
    letras = ""
    while indice:
        indice, resto = divmod(indice - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


def caminho_unico(caminho: Path) -> Path:
    if not caminho.exists():
        return caminho
    contador = 2
    while True:
        candidato = caminho.with_name(f"{caminho.stem} ({contador}){caminho.suffix}")
        if not candidato.exists():
            return candidato
        contador += 1
