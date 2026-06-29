from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config import carregar_config
from app.services.controle_fiscal_service import (
    ABA_PROCESSOS,
    ARQUIVO_CONTROLE,
    COLUNAS,
    STATUS_FISCAL_INICIAL,
    _normalizar_valor_chave,
)
from app.utils import limpar_espacos, parse_valor_brl


STATUS_HISTORICO = "APROVADO_HISTORICO"
COLUNAS_MANUAIS = {"Status Fiscal", "Data Envio Fiscal", "Observação", "ObservaÃ§Ã£o"}


@dataclass
class ResumoImportacao:
    arquivos_encontrados: int = 0
    novos_registros: int = 0
    duplicados: int = 0
    sem_pedido: int = 0
    sem_nf: int = 0
    sem_valor: int = 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa PDFs historicos de COMPLETOS para CONTROLE_FISCAL.xlsx.")
    parser.add_argument("--dry-run", action="store_true", help="Nao grava planilha; apenas mostra o resumo.")
    parser.add_argument("--ano", type=int, default=2026, help="Ano a importar. Padrao: 2026.")
    parser.add_argument("--mes", type=int, choices=range(1, 13), help="Mes numerico a importar.")
    args = parser.parse_args()

    config = carregar_config()
    base = Path(config.base_processamento_fiscal)
    resumo = importar_completos_controle(base, ano=args.ano, mes=args.mes, dry_run=args.dry_run)
    imprimir_resumo(resumo, dry_run=args.dry_run)


def importar_completos_controle(base: Path, ano: int = 2026, mes: int | None = None, dry_run: bool = False) -> ResumoImportacao:
    arquivos = listar_pdfs_completos(base, ano, mes)
    linhas = [montar_linha_historica(pdf) for pdf in arquivos]
    caminho_controle = base / "RELATORIOS" / ARQUIVO_CONTROLE
    resumo = ResumoImportacao(arquivos_encontrados=len(arquivos))

    chaves_existentes = carregar_chaves_existentes(caminho_controle)
    chaves_novas = set()
    for linha in linhas:
        if not linha.get("Pedido"):
            resumo.sem_pedido += 1
        if not linha.get("NF"):
            resumo.sem_nf += 1
        if linha.get("Valor") in (None, ""):
            resumo.sem_valor += 1

        chave = chave_historica(linha)
        if chave in chaves_existentes or chave in chaves_novas:
            resumo.duplicados += 1
            continue
        chaves_novas.add(chave)
        resumo.novos_registros += 1

    if not dry_run:
        gravar_linhas(caminho_controle, [linha for linha in linhas if chave_historica(linha) in chaves_novas])

    return resumo


def listar_pdfs_completos(base: Path, ano: int, mes: int | None = None) -> list[Path]:
    raiz_ano = base / "PROCESSOS" / str(ano)
    if not raiz_ano.exists():
        return []

    meses = [p for p in raiz_ano.iterdir() if p.is_dir()]
    if mes is not None:
        prefixo = f"{mes:02d}-"
        meses = [p for p in meses if p.name.startswith(prefixo)]

    arquivos: list[Path] = []
    for pasta_mes in sorted(meses):
        for pasta_dia in sorted(p for p in pasta_mes.iterdir() if p.is_dir()):
            completos = pasta_dia / "COMPLETOS"
            if completos.exists():
                arquivos.extend(sorted(completos.glob("*.pdf")))
    return arquivos


def montar_linha_historica(pdf: Path) -> dict:
    fornecedor, nf, valor = extrair_dados_nome(pdf.name)
    pedido = extrair_pedido_pdf(pdf)
    data_processamento = data_pasta_dia(pdf)
    return {
        "Data Processamento": data_processamento,
        "Pedido": pedido,
        "Fornecedor": fornecedor,
        "CNPJ": None,
        "NF": nf,
        "Valor": float(valor) if valor is not None else None,
        "Status Processo": STATUS_HISTORICO,
        "Arquivo Completo": str(pdf),
        "Status Fiscal": STATUS_FISCAL_INICIAL,
        "Data Envio Fiscal": None,
        "Observação": None,
    }


def extrair_dados_nome(nome: str) -> tuple[str | None, str | None, Decimal | None]:
    stem = Path(nome).stem
    partes = [limpar_espacos(p) for p in stem.split(" - ") if limpar_espacos(p)]
    fornecedor = partes[0] if partes else None

    nf = None
    for parte in partes[1:]:
        match = re.search(r"\b(?:NF|NFS|NFE|CTE|DACTE)\s*([A-Z0-9./-]+)", parte, flags=re.IGNORECASE)
        if match:
            nf = match.group(1).strip(" .-/")
            break

    valor = None
    valor_match = re.search(r"R\$\s*([\d.,]+)", stem, flags=re.IGNORECASE)
    if valor_match:
        valor = parse_valor_brl(valor_match.group(1))

    return fornecedor, nf, valor


def extrair_pedido_pdf(pdf: Path) -> str | None:
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(pdf))
        textos = []
        for page in reader.pages:
            textos.append(page.extract_text() or "")
        texto = "\n".join(textos)
    except Exception:
        return None

    padroes = [
        r"\bPEDIDO(?:\s+DE\s+COMPRA|\s+DO\s+CLIENTE)?\s*:?\s*(\d{3,})\b",
        r"\bPC\s*:?\s*(\d{3,})\b",
        r"\bP\.?C\.?\s*:?\s*(\d{3,})\b",
    ]
    pedidos = []
    for padrao in padroes:
        pedidos.extend(re.findall(padrao, texto, flags=re.IGNORECASE))
    unicos = sorted(set(pedidos))
    return ", ".join(unicos) if unicos else None


def data_pasta_dia(pdf: Path) -> str | None:
    try:
        data_ref = datetime.strptime(pdf.parent.parent.name, "%d-%m-%Y")
    except ValueError:
        return None
    return data_ref.strftime("%d/%m/%Y")


def carregar_chaves_existentes(caminho: Path) -> set[tuple[str, str, str, str, str]]:
    if not caminho.exists():
        return set()
    wb = load_workbook(caminho)
    if ABA_PROCESSOS not in wb.sheetnames:
        return set()
    ws = wb[ABA_PROCESSOS]
    indices = indices_colunas(ws)
    return {
        chave_historica({coluna: ws.cell(row=row, column=indices[coluna]).value for coluna in indices})
        for row in range(2, ws.max_row + 1)
    }


def gravar_linhas(caminho: Path, linhas: list[dict]) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(caminho) if caminho.exists() else Workbook()
    ws = obter_aba(wb)
    garantir_cabecalho(ws)
    indices = indices_colunas(ws)
    existentes = {
        chave_historica({coluna: ws.cell(row=row, column=indices[coluna]).value for coluna in indices}): row
        for row in range(2, ws.max_row + 1)
    }

    for linha in linhas:
        chave = chave_historica(linha)
        row_idx = existentes.get(chave)
        if row_idx:
            atualizar_linha(ws, row_idx, indices, linha)
            continue
        ws.append([linha.get(coluna) for coluna in COLUNAS])
        existentes[chave] = ws.max_row

    wb.save(caminho)


def obter_aba(wb):
    if ABA_PROCESSOS in wb.sheetnames:
        return wb[ABA_PROCESSOS]
    if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None:
        ws = wb.active
        ws.title = ABA_PROCESSOS
        return ws
    return wb.create_sheet(ABA_PROCESSOS)


def garantir_cabecalho(ws) -> None:
    atual = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    if not any(atual):
        for col, coluna in enumerate(COLUNAS, start=1):
            ws.cell(row=1, column=col, value=coluna)
        return
    for coluna in COLUNAS:
        if coluna not in atual:
            ws.cell(row=1, column=len(atual) + 1, value=coluna)
            atual.append(coluna)


def indices_colunas(ws) -> dict[str, int]:
    indices = {
        str(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=col).value
    }
    if "ObservaÃ§Ã£o" in indices and "Observação" not in indices:
        indices["Observação"] = indices["ObservaÃ§Ã£o"]
    return indices


def atualizar_linha(ws, row_idx: int, indices: dict[str, int], linha: dict) -> None:
    for coluna, valor in linha.items():
        if coluna in COLUNAS_MANUAIS:
            continue
        if coluna in indices:
            ws.cell(row=row_idx, column=indices[coluna], value=valor)
    status_col = indices.get("Status Fiscal")
    if status_col and ws.cell(row=row_idx, column=status_col).value in (None, ""):
        ws.cell(row=row_idx, column=status_col, value=STATUS_FISCAL_INICIAL)


def chave_historica(linha: dict) -> tuple[str, str, str, str, str]:
    return (
        normalizar_chave(linha.get("Pedido")),
        normalizar_chave(linha.get("NF")),
        normalizar_chave(linha.get("Fornecedor")),
        _normalizar_valor_chave(linha.get("Valor")),
        normalizar_chave(linha.get("Arquivo Completo")),
    )


def normalizar_chave(valor) -> str:
    return " ".join(str(valor or "").strip().upper().split())


def imprimir_resumo(resumo: ResumoImportacao, dry_run: bool) -> None:
    print(f"Modo: {'DRY-RUN' if dry_run else 'EXECUCAO'}")
    print(f"Arquivos encontrados: {resumo.arquivos_encontrados}")
    print(f"Novos registros: {resumo.novos_registros}")
    print(f"Duplicados: {resumo.duplicados}")
    print(f"Sem pedido: {resumo.sem_pedido}")
    print(f"Sem NF: {resumo.sem_nf}")
    print(f"Sem valor: {resumo.sem_valor}")


if __name__ == "__main__":
    main()
