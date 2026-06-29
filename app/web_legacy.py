from __future__ import annotations

import json
import hashlib
import logging
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from flask import Flask, abort, jsonify, redirect, render_template, request, send_file, url_for

from .config import ROOT_DIR, carregar_config
from .agrupador import agrupar_documentos
from .classificador import classificar_texto
from .extrator_campos import extrair_campos
from .leitor_pdf import extrair_texto_paginas
from .main import listar_pdfs, normalizar_paginas_mesmo_arquivo, processar_aprovado
from .paths import montar_pastas
from .renomeador import limpar_nome_fornecedor
from .utils import normalizar_texto
from .validador import validar_processo
from .services.execucao_service import carregar_ultima_execucao


app = Flask(
    __name__,
    template_folder=str(ROOT_DIR / "templates"),
    static_folder=str(ROOT_DIR / "static"),
)
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
APROVACOES_MANUAIS = ROOT_DIR / "config" / "aprovacoes_manuais.json"


@app.after_request
def sem_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.context_processor
def assets_versionados():
    def asset_url(caminho: str) -> str:
        arquivo = ROOT_DIR / "static" / caminho
        versao = int(arquivo.stat().st_mtime) if arquivo.exists() else 1
        return url_for("static", filename=caminho, v=versao)

    return {"asset_url": asset_url}


def relatorios_dir() -> Path:
    return montar_pastas(carregar_config()).relatorios


def logs_dir() -> Path:
    return montar_pastas(carregar_config()).logs


def arquivo_mais_recente(padrao: str) -> Path | None:
    arquivos = sorted(relatorios_dir().glob(padrao), key=lambda p: p.stat().st_mtime, reverse=True)
    return arquivos[0] if arquivos else None


def carregar_linhas_relatorio() -> tuple[list[dict], Path | None]:
    caminho = arquivo_mais_recente("relatorio_processamento_*.xlsx")
    if not caminho:
        return [], None

    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], caminho

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return [], caminho

    cabecalho = [str(coluna or "").strip() for coluna in linhas[0]]
    dados = []
    for indice, valores in enumerate(linhas[1:], start=1):
        linha = {cabecalho[i]: valores[i] if i < len(valores) else None for i in range(len(cabecalho))}
        linha["_id"] = indice
        linha["_confianca_pct"] = calcular_confianca(linha)
        linha["_diferenca"] = diferenca_linha(linha)
        dados.append(linha)
    return dados, caminho


def carregar_processos_txt() -> tuple[list[dict], dict, Path | None]:
    caminho = arquivo_mais_recente("resumo_processamento_*.txt")
    if not caminho:
        return [], {"aprovados": 0, "pendentes": 0}, None

    texto = caminho.read_text(encoding="utf-8", errors="replace")
    resumo = {
        "aprovados": extrair_inteiro(texto, r"Processos aprovados:\s*(\d+)"),
        "pendentes": extrair_inteiro(texto, r"Processos pendentes:\s*(\d+)"),
    }
    processos = []
    atual: dict | None = None
    for linha in texto.splitlines():
        cabecalho = re.match(r"^(PROCESSO_\d+)\s+-\s+(.+)$", linha.strip())
        if cabecalho:
            atual = {
                "id": cabecalho.group(1),
                "status": cabecalho.group(2).strip(),
                "erros": [],
                "documentos": [],
            }
            processos.append(atual)
            continue
        if not atual:
            continue
        if linha.startswith("Erros:"):
            atual["erros"] = [erro.strip() for erro in linha.replace("Erros:", "", 1).split(";") if erro.strip()]
            continue
        doc = re.match(r"^\s+-\s+(.+?)\s+\|\s+(.+?)\s+\|\s+pagina\s+(\d+)", linha)
        if doc:
            atual["documentos"].append(
                {
                    "tipo": doc.group(1).strip(),
                    "arquivo": doc.group(2).strip(),
                    "pagina": doc.group(3).strip(),
                }
            )
    return processos, resumo, caminho


def extrair_inteiro(texto: str, padrao: str) -> int:
    match = re.search(padrao, texto)
    return int(match.group(1)) if match else 0


def parse_brl(valor) -> Decimal | None:
    if valor in (None, ""):
        return None
    texto = str(valor).replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return None


def formatar_brl(valor: Decimal | None) -> str:
    if valor is None:
        return "-"
    texto = f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {texto}"


def diferenca_linha(linha: dict) -> str:
    nf = parse_brl(linha.get("Valor NF"))
    boleto = parse_brl(linha.get("Valor Boleto(s)"))
    pedido = parse_brl(linha.get("Valor Pedido(s)"))
    referencia = boleto if boleto is not None else pedido
    if nf is None or referencia is None:
        return "-"
    return formatar_brl(nf - referencia)


def calcular_confianca(linha: dict) -> int:
    valor = linha.get("Confianca OCR")
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return 92 if linha.get("Status") == "APROVADO" else 68
    if numero <= 1:
        numero *= 100
    return max(0, min(100, round(numero)))


def montar_processos_operacionais(linhas: list[dict], processos_txt: list[dict]) -> list[dict]:
    aprovacoes = carregar_aprovacoes_manuais()
    por_arquivo: dict[str, list[dict]] = defaultdict(list)
    for linha in linhas:
        if linha.get("Arquivo Original"):
            por_arquivo[str(linha["Arquivo Original"])].append(linha)

    processos = []
    for processo in processos_txt:
        nomes = {doc["arquivo"] for doc in processo.get("documentos", [])}
        linhas_processo = [linha for nome in nomes for linha in por_arquivo.get(nome, [])]
        docs_txt = processo.get("documentos", [])
        tipos = {doc.get("tipo") for doc in docs_txt}
        fornecedor = fornecedor_processo(linhas_processo) or "Fornecedor nao identificado"
        nf = limpar_nf_ui(campo_por_tipo(linhas_processo, "NF", ["NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE", "BOLETO", "PEDIDO_COMPRA"]))
        pedido = campo_por_tipo(linhas_processo, "Pedido(s)", ["PEDIDO_COMPRA", "NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE", "BOLETO"])
        valor_nf = campo_por_tipo(linhas_processo, "Valor NF", ["NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE", "BOLETO", "PEDIDO_COMPRA"])
        valor_produtos_nf = campo_por_tipo(linhas_processo, "Valor Produtos NF", ["NF_PRODUTO"])
        valor_boleto = campo_por_tipo(linhas_processo, "Valor Boleto(s)", ["BOLETO", "NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE", "PEDIDO_COMPRA"])
        valor_pedido = campo_por_tipo(linhas_processo, "Valor Pedido(s)", ["PEDIDO_COMPRA", "NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE", "BOLETO"])
        status = status_operacional(processo.get("status"), processo.get("erros", []))
        documentos = resumo_documentos(tipos)
        confianca = min([linha.get("_confianca_pct", 100) for linha in linhas_processo], default=100)
        confianca_geral = primeiro_numero([linha.get("Confianca Processo") for linha in linhas_processo], confianca)
        processo_operacional = {
                "id": processo["id"],
                "status": status,
                "status_raw": processo.get("status"),
                "fornecedor": fornecedor,
                "cnpj": primeiro_texto([linha.get("CNPJ") or linha.get("CNPJ Fornecedor") for linha in linhas_processo]),
                "nf": nf or "-",
                "pedido": pedido or "-",
                "valor": valor_nf or valor_boleto or valor_pedido or "R$ 0,00",
                "valor_nf": valor_nf or "-",
                "valor_produtos_nf": valor_produtos_nf or "-",
                "valor_boleto": valor_boleto or "-",
                "valor_pedido": valor_pedido or "-",
                "diferenca": diferenca_processo(valor_nf, valor_boleto, valor_pedido),
                "data": primeiro_texto([linha.get("Data") or linha.get("Data Processamento") for linha in linhas_processo]) or "-",
                "documentos": documentos,
                "documentos_label": f"{documentos['presentes']}/{documentos['esperados']}",
                "legendas": documentos["legendas"],
                "conferencia": confianca,
                "confianca_geral": confianca_geral,
                "proxima_acao": proxima_acao(status),
                "responsavel": responsavel(status),
                "erros": processo.get("erros", []),
                "pedido_validacao": status_validacao_pedido(processo.get("erros", [])),
                "linhas": linhas_processo,
                "docs": docs_txt,
                "aprovacao_manual": aprovacoes.get(chave_aprovacao_docs(docs_txt)),
            }
        if processo_operacional["aprovacao_manual"] and not aprovacao_manual_valida(
            processo_operacional,
            processo_operacional["aprovacao_manual"],
        ):
            processo_operacional["aprovacao_manual"] = None
        if processo_operacional["aprovacao_manual"]:
            processo_operacional["status"] = "FINALIZADO"
            processo_operacional["proxima_acao"] = "Processo finalizado"
        processos.append(processo_operacional)
    return processos


def chave_aprovacao_docs(docs: list[dict]) -> str:
    nomes = sorted(
        {
            str(doc.get("arquivo") or doc.get("arquivo_nome") or "")
            for doc in docs
            if doc.get("arquivo") or doc.get("arquivo_nome")
        }
    )
    return hashlib.sha256("|".join(nomes).encode("utf-8")).hexdigest()


def carregar_aprovacoes_manuais() -> dict:
    if not APROVACOES_MANUAIS.exists():
        return {}
    try:
        return json.loads(APROVACOES_MANUAIS.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}


def salvar_aprovacoes_manuais(aprovacoes: dict) -> None:
    APROVACOES_MANUAIS.parent.mkdir(parents=True, exist_ok=True)
    temporario = APROVACOES_MANUAIS.with_suffix(".tmp")
    temporario.write_text(json.dumps(aprovacoes, ensure_ascii=False, indent=2), encoding="utf-8")
    temporario.replace(APROVACOES_MANUAIS)


def aprovacao_manual_valida(processo: dict, aprovacao: dict) -> bool:
    arquivo = Path(str(aprovacao.get("arquivo_unido") or ""))
    return arquivo_corresponde_processo(arquivo, processo)


def primeiro_texto(valores) -> str | None:
    for valor in valores:
        if valor not in (None, "", "R$ 0,00"):
            return str(valor)
    return None


def primeiro_numero(valores, padrao: int = 0) -> int:
    for valor in valores:
        try:
            return max(0, min(100, round(float(valor))))
        except (TypeError, ValueError):
            continue
    return padrao


def campo_por_tipo(linhas: list[dict], campo: str, ordem_tipos: list[str]) -> str | None:
    for tipo in ordem_tipos:
        valor = primeiro_texto(linha.get(campo) for linha in linhas if linha.get("Tipo Documento") == tipo)
        if valor:
            return valor
    return primeiro_texto(linha.get(campo) for linha in linhas)


def limpar_nf_ui(valor: str | None) -> str | None:
    if valor in (None, ""):
        return None
    texto = str(valor).strip()
    return None if texto in {"0", "1", "2"} else texto


def fornecedor_processo(linhas: list[dict]) -> str | None:
    ordem = ["PEDIDO_COMPRA", "NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE", "BOLETO"]
    for tipo in ordem:
        for linha in linhas:
            if linha.get("Tipo Documento") != tipo:
                continue
            nome = limpar_fornecedor_ui(str(linha.get("Fornecedor") or ""))
            if nome:
                return nome
    return None


def limpar_fornecedor_ui(nome: str) -> str | None:
    nome = limpar_nome_fornecedor(nome or "")
    nome = re.sub(r"\s*\|\s*", " ", nome)
    nome = re.sub(r"\s+", " ", nome).strip(" -")
    nome_norm = nome.upper()
    bloqueados = [
        "DOCUMENTO AUXILIAR",
        "DATA DO DOCUMENTO",
        "SEFAZ AUTORIZADORA",
        "E-MAIL:",
        "CNPJ ",
        "CPFSCNPI",
        "CHAVE DE ACESSO",
    ]
    if not nome or any(bloqueado in nome_norm for bloqueado in bloqueados):
        return None
    partes = [parte.strip() for parte in nome.split(" - ") if parte.strip()]
    if len(partes) >= 2 and partes[-1].upper() in partes[0].upper():
        nome = partes[0]
    return nome[:70]


def status_operacional(status: str | None, erros: list[str]) -> str:
    if status == "APROVADO":
        return "APROVADO"
    texto = " ".join(erros).upper()
    if "NOTA FISCAL" in texto or "SEM NOTA" in texto:
        return "AGUARDANDO NF"
    if "BOLETO" in texto:
        return "AGUARDANDO BOLETO"
    if "PEDIDO DE COMPRA NAO IDENTIFICADO" in texto or "SEM PEDIDO" in texto:
        return "AGUARDANDO PEDIDO"
    if "PEDIDO" in texto and ("DIVERGE" in texto or "DIVERGENTE" in texto):
        return "PEDIDO DIVERGENTE"
    if "PEDIDO" in texto:
        return "REVISAR PEDIDO"
    if "OCR" in texto or "CLASSIFICACAO" in texto:
        return "PROCESSANDO OCR"
    return "PENDENTE"


def status_validacao_pedido(erros: list[str]) -> str:
    texto = " ".join(erros).upper()
    if "PEDIDO DE COMPRA NAO IDENTIFICADO" in texto or "SEM PEDIDO" in texto:
        return "Aguardando"
    if "PEDIDO" in texto and ("DIVERGE" in texto or "DIVERGENTE" in texto):
        return "Divergente"
    if "PEDIDO" in texto:
        return "Revisar"
    return "OK"


def resumo_documentos(tipos: set[str]) -> dict:
    tem_nf = bool(tipos & {"NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO"})
    tem_cte = "DACTE" in tipos
    tem_boleto = "BOLETO" in tipos
    tem_pedido = "PEDIDO_COMPRA" in tipos
    if tem_cte:
        esperados = {"NF": tem_cte, "BOLETO": tem_boleto}
    else:
        esperados = {"NF": tem_nf, "BOLETO": tem_boleto, "PEDIDO": tem_pedido}
    return {
        "presentes": sum(1 for ok in esperados.values() if ok),
        "esperados": len(esperados),
        "legendas": [{"nome": nome, "ok": ok} for nome, ok in esperados.items()],
    }


def diferenca_processo(valor_nf, valor_boleto, valor_pedido) -> str:
    nf = parse_brl(valor_nf)
    referencia = parse_brl(valor_boleto) or parse_brl(valor_pedido)
    if nf is None or referencia is None:
        return "-"
    return formatar_brl(nf - referencia)


def proxima_acao(status: str) -> str:
    acoes = {
        "APROVADO": "Enviar ZapSign",
        "AGUARDANDO NF": "Anexar nota fiscal",
        "AGUARDANDO BOLETO": "Anexar boleto",
        "AGUARDANDO PEDIDO": "Anexar pedido",
        "PEDIDO DIVERGENTE": "Conferir numero do pedido",
        "REVISAR PEDIDO": "Conferir pedido",
        "PROCESSANDO OCR": "Revisar OCR",
        "ENVIADO ZAPSIGN": "Aguardar assinatura",
        "ASSINADO": "Finalizar processo",
    }
    return acoes.get(status, "Conferir pendencia")


def responsavel(status: str) -> str:
    if status in {"APROVADO", "ENVIADO ZAPSIGN", "ASSINADO"}:
        return "Financeiro"
    if status == "PROCESSANDO OCR":
        return "Tecnico"
    return "Conferencia"


def montar_kpis(processos: list[dict], linhas: list[dict]) -> dict:
    total_nf = Decimal("0")
    for processo in processos:
        if processo["status"] in {"APROVADO", "FINALIZADO"}:
            total_nf += parse_brl(processo.get("valor")) or Decimal("0")
    aprovados = sum(1 for processo in processos if processo["status"] in {"APROVADO", "FINALIZADO"})
    pendentes = sum(1 for processo in processos if processo["status"] not in {"APROVADO", "FINALIZADO"})
    arquivos = {linha.get("Arquivo Original") for linha in linhas if linha.get("Arquivo Original")}
    erros = sum(1 for linha in linhas if linha.get("Erro Encontrado"))
    ocr = sum(1 for linha in linhas if linha.get("_confianca_pct", 100) < 90)
    total = max(1, len(processos))
    confianca_media = round(sum(p.get("confianca_geral", p.get("conferencia", 0)) for p in processos) / total)
    return {
        "arquivos": len(arquivos),
        "processos_dia": len(processos),
        "aprovados": aprovados,
        "pendencias": pendentes,
        "erros": erros,
        "valor_total": formatar_brl(total_nf),
        "ocr": ocr,
        "aprovados_pct": f"{(aprovados / total) * 100:.2f}",
        "pendencias_pct": f"{(pendentes / total) * 100:.2f}",
        "erros_pct": f"{(erros / total) * 100:.2f}",
        "aguardando_assinatura": aprovados,
        "zapsign": aprovados,
        "confianca_media": confianca_media,
    }


def montar_graficos(processos: list[dict], resumo_txt: dict) -> dict:
    status = {
        "aprovados": resumo_txt.get("aprovados", 0),
        "pendentes": resumo_txt.get("pendentes", 0),
    }
    total_status = max(1, status["aprovados"] + status["pendentes"])
    status["aprovados_pct"] = int((status["aprovados"] / total_status) * 100)
    por_fornecedor: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for processo in processos:
        if processo.get("status") not in {"APROVADO", "FINALIZADO"}:
            continue
        fornecedor = str(processo.get("fornecedor") or "Nao identificado")[:42]
        por_fornecedor[fornecedor] += parse_brl(processo.get("valor")) or Decimal("0")
    top = sorted(por_fornecedor.items(), key=lambda item: item[1], reverse=True)[:10]
    maior = max([valor for _, valor in top], default=Decimal("1"))
    fornecedores = [
        {"nome": nome, "valor": formatar_brl(valor), "pct": int((valor / maior) * 100) if maior else 0}
        for nome, valor in top
    ]
    pendencias = Counter()
    for processo in processos:
        texto = " ".join(processo.get("erros", [])).upper()
        if "NOTA FISCAL" in texto or "SEM NOTA" in texto:
            pendencias["Falta NF"] += 1
        if "PEDIDO" in texto and ("NAO IDENTIFICADO" in texto or "SEM ANEXO" in texto or "SEM PEDIDO" in texto):
            pendencias["Falta Pedido"] += 1
        if "BOLETO" in texto:
            pendencias["Falta Boleto"] += 1
        if "VALOR" in texto or "DIVERGE" in texto:
            pendencias["Divergencia de Valor"] += 1
        if "CNPJ" in texto:
            pendencias["Divergencia de CNPJ"] += 1
        if processo.get("conferencia", 100) < 70:
            pendencias["OCR Baixa Confianca"] += 1
    documentos = Counter(
        str(linha.get("Tipo Documento") or "DESCONHECIDO")
        for processo in processos
        for linha in processo.get("linhas", [])
    )
    return {
        "status": status,
        "fornecedores": fornecedores,
        "pendencias": [{"nome": nome, "total": total} for nome, total in pendencias.most_common()],
        "documentos": [{"nome": nome, "total": total} for nome, total in documentos.most_common()],
    }


def carregar_logs() -> tuple[list[dict], Path | None]:
    arquivos = sorted(logs_dir().glob("robo_fiscal_*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not arquivos:
        return [], None
    caminho = arquivos[0]
    linhas = caminho.read_text(encoding="utf-8", errors="replace").splitlines()[-220:]
    eventos = []
    for linha in linhas:
        partes = [parte.strip() for parte in linha.split("|", 2)]
        if len(partes) == 3:
            eventos.append({"hora": partes[0], "nivel": partes[1], "mensagem": partes[2]})
        else:
            eventos.append({"hora": "", "nivel": "INFO", "mensagem": linha})
    return eventos, caminho


def contexto_base() -> dict:
    linhas, relatorio_xlsx = carregar_linhas_relatorio()
    processos, resumo_txt, relatorio_txt = carregar_processos_txt()
    processos_operacionais = montar_processos_operacionais(linhas, processos)
    logs, log_atual = carregar_logs()
    pastas = montar_pastas(carregar_config())
    ultima_execucao = carregar_ultima_execucao(pastas.logs)
    duracao = float(ultima_execucao.get("duracao_segundos") or 0)
    ultima_execucao["duracao_formatada"] = f"{int(duracao // 60):02d}:{int(duracao % 60):02d}"
    return {
        "linhas": linhas,
        "processos": processos,
        "processos_operacionais": processos_operacionais,
        "kpis": montar_kpis(processos_operacionais, linhas),
        "graficos": montar_graficos(processos_operacionais, resumo_txt),
        "relatorio_xlsx": relatorio_xlsx,
        "relatorio_txt": relatorio_txt,
        "logs": logs,
        "log_atual": log_atual,
        "config": carregar_config(),
        "pastas": pastas,
        "ultima_execucao": ultima_execucao,
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }


@app.get("/")
def dashboard():
    contexto = contexto_base()
    contexto["active"] = "dashboard"
    return render_template("dashboard.html", **contexto)


@app.get("/processos")
def processos():
    contexto = contexto_base()
    contexto["active"] = "processos"
    return render_template("processos.html", **contexto)


@app.get("/assinaturas")
def assinaturas():
    contexto = contexto_base()
    contexto["active"] = "assinaturas"
    contexto["processos_assinatura"] = [
        processo for processo in contexto["processos_operacionais"] if processo["status"] == "APROVADO"
    ]
    return render_template("assinaturas.html", **contexto)


@app.get("/pendencias")
def pendencias():
    contexto = contexto_base()
    contexto["active"] = "pendencias"
    contexto["linhas"] = [linha for linha in contexto["linhas"] if linha.get("Status") != "APROVADO" or linha.get("Erro Encontrado")]
    contexto["processos"] = [p for p in contexto["processos"] if p.get("status") != "APROVADO"]
    contexto["processos_operacionais"] = [p for p in contexto["processos_operacionais"] if p.get("status") != "APROVADO"]
    return render_template("pendencias.html", **contexto)


@app.get("/fornecedores")
def fornecedores():
    contexto = contexto_base()
    contexto["active"] = "fornecedores"
    fornecedores_map: dict[str, dict] = {}
    for processo in contexto["processos_operacionais"]:
        item = fornecedores_map.setdefault(
            processo["fornecedor"],
            {"nome": processo["fornecedor"], "processos": 0, "aprovados": 0, "valor": Decimal("0")},
        )
        item["processos"] += 1
        if processo["status"] in {"APROVADO", "FINALIZADO"}:
            item["aprovados"] += 1
            item["valor"] += parse_brl(processo["valor"]) or Decimal("0")
    contexto["fornecedores"] = [
        {**item, "valor_fmt": formatar_brl(item["valor"])}
        for item in sorted(fornecedores_map.values(), key=lambda f: f["processos"], reverse=True)
    ]
    return render_template("fornecedores.html", **contexto)


@app.get("/relatorios")
def relatorios():
    contexto = contexto_base()
    contexto["active"] = "relatorios"
    contexto["arquivos_relatorio"] = sorted(relatorios_dir().glob("*.*"), key=lambda p: p.stat().st_mtime, reverse=True)
    return render_template("relatorios.html", **contexto)


@app.get("/logs")
def logs():
    contexto = contexto_base()
    contexto["active"] = "logs_tecnicos"
    return render_template("logs.html", **contexto)


@app.get("/configuracoes")
def configuracoes():
    contexto = contexto_base()
    contexto["active"] = "configuracoes"
    return render_template("configuracoes.html", **contexto)


@app.get("/processo/<processo_id>")
def detalhe_processo(processo_id: str):
    contexto = contexto_base()
    processo = next((p for p in contexto["processos_operacionais"] if p["id"] == processo_id), None)
    if not processo:
        abort(404)
    contexto["processo"] = processo
    contexto["linhas_processo"] = processo["linhas"]
    contexto["pdf_preview"] = localizar_pdf_preview(contexto["linhas_processo"])
    contexto["pdf_tabs"] = montar_pdf_tabs(contexto["linhas_processo"], processo)
    contexto["active"] = "processos"
    return render_template("detalhe.html", **contexto)


@app.post("/api/processo/<processo_id>/aprovar")
def aprovar_processo(processo_id: str):
    contexto = contexto_base()
    processo_ui = next((p for p in contexto["processos_operacionais"] if p["id"] == processo_id), None)
    if not processo_ui:
        return jsonify({"ok": False, "erro": "Processo nao localizado no relatorio atual."}), 404
    if processo_ui.get("aprovacao_manual"):
        return jsonify({"ok": True, "redirect": url_for("detalhe_processo", processo_id=processo_id)})

    config = carregar_config()
    pastas = montar_pastas(config)
    nomes_alvo = {doc.get("arquivo") for doc in processo_ui.get("docs", []) if doc.get("arquivo")}
    try:
        arquivo_existente = localizar_processo_completo(processo_ui)
        if arquivo_existente:
            registrar_aprovacao_manual(processo_id, processo_ui, arquivo_existente)
            return jsonify(
                {
                    "ok": True,
                    "arquivo": str(arquivo_existente),
                    "recuperado": True,
                    "redirect": url_for("detalhe_processo", processo_id=processo_id),
                }
            )

        processos = reconstruir_processos_entrada(config, pastas)
        processo = localizar_processo_para_aprovacao(processos, nomes_alvo)
        if not processo:
            return jsonify({"ok": False, "erro": "Nao foi possivel reconstruir este processo a partir da pasta ENTRADA."}), 409

        processo["status"] = "APROVADO_MANUAL"
        processo["aprovacao_manual"] = True
        processar_aprovado(processo, pastas)
        arquivo_unido = Path(processo["arquivo_unido"])
        if not arquivo_unido.exists() or arquivo_unido.stat().st_size == 0:
            raise RuntimeError("O PDF final nao foi gerado corretamente.")

        registrar_aprovacao_manual(processo_id, processo_ui, arquivo_unido)
        return jsonify({"ok": True, "arquivo": str(arquivo_unido), "redirect": url_for("detalhe_processo", processo_id=processo_id)})
    except Exception as exc:
        logging.exception("Falha ao aprovar manualmente %s: %s", processo_id, exc)
        return jsonify({"ok": False, "erro": str(exc)}), 500


def registrar_aprovacao_manual(processo_id: str, processo_ui: dict, arquivo_unido: Path) -> None:
    aprovacoes = carregar_aprovacoes_manuais()
    chave = chave_aprovacao_docs(processo_ui.get("docs", []))
    aprovacoes[chave] = {
        "processo_id": processo_id,
        "aprovado_em": datetime.now().isoformat(timespec="seconds"),
        "arquivo_unido": str(arquivo_unido),
        "pendencias_ignoradas": processo_ui.get("erros", []),
    }
    salvar_aprovacoes_manuais(aprovacoes)


def reconstruir_processos_entrada(config, pastas) -> list[dict]:
    documentos = []
    for pdf in listar_pdfs(pastas.entrada):
        for pagina in extrair_texto_paginas(pdf, config.limite_texto_minimo_pdf_digital, config.tesseract_path):
            tipo, confianca = classificar_texto(pagina["texto"])
            documentos.append(extrair_campos(pagina, tipo, confianca))
    documentos = normalizar_paginas_mesmo_arquivo(documentos)
    return [validar_processo(processo, config.tolerancia_valor) for processo in agrupar_documentos(documentos)]


def localizar_processo_para_aprovacao(processos: list[dict], nomes_alvo: set[str]) -> dict | None:
    candidatos = []
    for processo in processos:
        nomes = {Path(str(doc.get("arquivo_origem") or "")).name for doc in processo.get("documentos", [])}
        intersecao = len(nomes & nomes_alvo)
        if intersecao:
            candidatos.append((intersecao, len(nomes_alvo - nomes), processo))
    if not candidatos:
        return None
    candidatos.sort(key=lambda item: (item[0], -item[1]), reverse=True)
    melhor = candidatos[0]
    return melhor[2] if melhor[0] == len(nomes_alvo) else None


@app.post("/api/run")
def api_run():
    payload = request.get_json(silent=True) or {}
    dry_run = request.form.get("dry_run") == "1" or payload.get("dry_run") is True
    comando = [sys.executable, "-B", "-m", "app.main"]
    if dry_run:
        comando.append("--dry-run")
    resultado = subprocess.run(
        comando,
        cwd=str(ROOT_DIR),
        text=True,
        capture_output=True,
        timeout=1200,
        encoding="utf-8",
        errors="replace",
    )
    return jsonify(
        {
            "ok": resultado.returncode == 0,
            "returncode": resultado.returncode,
            "stdout": resultado.stdout,
            "stderr": resultado.stderr,
            "redirect": url_for("dashboard"),
        }
    )


@app.get("/pdf")
def pdf():
    caminho = request.args.get("path", "")
    arquivo = caminho_pdf_seguro(caminho)
    if not arquivo:
        abort(404)
    return send_file(arquivo, mimetype="application/pdf", as_attachment=False, download_name=arquivo.name)


@app.get("/pdf/view")
def pdf_viewer():
    caminho = request.args.get("path", "")
    arquivo = caminho_pdf_seguro(caminho)
    if not arquivo:
        abort(404)

    try:
        import fitz

        with fitz.open(arquivo) as documento:
            total_paginas = documento.page_count
    except (ImportError, OSError, RuntimeError, ValueError):
        abort(422)

    paginas = [
        url_for("pdf_page", path=str(arquivo), page=numero)
        for numero in range(total_paginas)
    ]
    return render_template(
        "pdf_viewer.html",
        arquivo=arquivo,
        paginas=paginas,
        pdf_original=url_for("pdf", path=str(arquivo)),
    )


@app.get("/pdf/page/<int:page>")
def pdf_page(page: int):
    caminho = request.args.get("path", "")
    arquivo = caminho_pdf_seguro(caminho)
    if not arquivo:
        abort(404)

    try:
        import fitz

        with fitz.open(arquivo) as documento:
            if page < 0 or page >= documento.page_count:
                abort(404)
            pagina = documento.load_page(page)
            imagem = pagina.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False).tobytes("png")
    except (ImportError, OSError, RuntimeError, ValueError):
        abort(422)

    return send_file(BytesIO(imagem), mimetype="image/png", max_age=3600)


def localizar_pdf_preview(linhas: list[dict]) -> str | None:
    for linha in linhas:
        arquivo = primeiro_pdf_encontrado(linha)
        if arquivo:
            return url_for("pdf_viewer", path=str(arquivo))
    return None


def montar_pdf_tabs(linhas: list[dict], processo: dict | None = None) -> list[dict]:
    grupos = [
        ("Nota Fiscal", {"NF_PRODUTO", "NF_SERVICO", "RECIBO_LOCACAO", "DACTE"}),
        ("Boleto", {"BOLETO"}),
        ("Pedido", {"PEDIDO_COMPRA"}),
        ("PDF Final", {"PDF_FINAL"}),
    ]
    tabs = []
    for label, tipos in grupos:
        arquivos = []
        if "PDF_FINAL" in tipos:
            arquivo = primeiro_pdf_final(linhas, processo)
            if arquivo:
                arquivos.append(arquivo)
        else:
            vistos = set()
            for linha in linhas:
                if linha.get("Tipo Documento") not in tipos:
                    continue
                arquivo = primeiro_pdf_encontrado(linha)
                if not arquivo or str(arquivo) in vistos:
                    continue
                vistos.add(str(arquivo))
                arquivos.append(arquivo)
        anexos = [
            {
                "nome": arquivo.name,
                "url": url_for("pdf_viewer", path=str(arquivo)),
            }
            for arquivo in arquivos
        ]
        tabs.append({"label": label, "url": anexos[0]["url"] if anexos else None, "anexos": anexos})
    return tabs


def primeiro_pdf_encontrado(linha: dict | None) -> Path | None:
    if not linha:
        return None
    for campo in ("Arquivo Final", "Arquivo Original"):
        nome = linha.get(campo)
        if not nome:
            continue
        arquivo = caminho_pdf_seguro(str(nome))
        if arquivo:
            return arquivo
    return None


def primeiro_pdf_final(linhas: list[dict], processo: dict | None = None) -> Path | None:
    completo = localizar_processo_completo(processo) if processo else None
    if completo:
        return completo
    for linha in linhas:
        nome = linha.get("Arquivo Final")
        if not nome:
            continue
        arquivo = caminho_pdf_seguro(str(nome))
        if arquivo:
            return arquivo
    return None


def localizar_processo_completo(processo: dict | None) -> Path | None:
    if not processo:
        return None
    config = carregar_config()
    pastas = montar_pastas(config)
    pasta = pastas.completos
    if not pasta.exists():
        return None

    fornecedor = normalizar_busca(processo.get("fornecedor"))
    nf = normalizar_busca(processo.get("nf"))
    valor = normalizar_busca(processo.get("valor"))
    candidatos = []
    for arquivo in pasta.glob("*.pdf"):
        if not arquivo_corresponde_processo(arquivo, processo):
            continue
        nome = normalizar_busca(arquivo.name)
        score = 0
        palavras_fornecedor = [p for p in fornecedor.split() if len(p) >= 4][:4]
        score += sum(2 for palavra in palavras_fornecedor if palavra in nome)
        if nf and nf != "-" and nf in nome:
            score += 5
        if valor and valor.replace("R", "")[:6] in nome:
            score += 2
        if "PROCESSO COMPLETO" in nome:
            score += 2
        if score:
            candidatos.append((score, arquivo.stat().st_mtime, arquivo))
    if not candidatos:
        return None
    candidatos.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return candidatos[0][2]


def arquivo_corresponde_processo(arquivo: Path, processo: dict) -> bool:
    if not arquivo.exists() or not arquivo.is_file() or arquivo.suffix.lower() != ".pdf" or arquivo.stat().st_size == 0:
        return False
    nome = normalizar_busca(arquivo.name)
    nf = normalizar_busca(processo.get("nf"))
    fornecedor = normalizar_busca(processo.get("fornecedor"))
    palavras_fornecedor = [
        palavra
        for palavra in fornecedor.split()
        if len(palavra) >= 4 and palavra not in {"LTDA", "INDUSTRIA", "COMERCIO", "SERVICOS"}
    ][:6]
    fornecedor_confere = any(palavra in nome for palavra in palavras_fornecedor)
    if nf and nf != "-":
        nf_confere = bool(re.search(rf"\b{re.escape(nf)}\b", nome))
        return nf_confere and fornecedor_confere
    valor = normalizar_busca(processo.get("valor"))
    valor_confere = bool(valor and valor.replace("R", "")[:6] in nome)
    return fornecedor_confere and valor_confere


def normalizar_busca(valor: str | None) -> str:
    texto = normalizar_texto(str(valor or ""))
    return re.sub(r"[^A-Z0-9]+", " ", texto).strip()


def caminho_pdf_seguro(caminho: str) -> Path | None:
    config = carregar_config()
    pastas = montar_pastas(config)
    bases = [
        pastas.entrada,
        pastas.processados,
        pastas.completos,
        pastas.pendencias_dia,
        pastas.pendencias_geral,
        pastas.assinados,
    ]
    candidato = Path(caminho)
    candidatos = [candidato] if candidato.is_absolute() else [base / candidato for base in bases]
    for item in candidatos:
        try:
            resolvido = item.resolve()
        except OSError:
            continue
        if not resolvido.exists() or resolvido.suffix.lower() != ".pdf":
            continue
        for base in bases:
            try:
                resolvido.relative_to(base.resolve())
                return resolvido
            except ValueError:
                continue
    return None


@app.template_filter("brl")
def filtro_brl(valor):
    return formatar_brl(parse_brl(valor))


@app.template_filter("data_path")
def filtro_data_path(caminho: Path | None):
    if not caminho:
        return "-"
    try:
        modificado = datetime.fromtimestamp(caminho.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
    except OSError:
        return "-"
    return modificado


if __name__ == "__main__":
    porta = int(os.environ.get("PORT", "5000"))
    app.run(host="127.0.0.1", port=porta, debug=False)
