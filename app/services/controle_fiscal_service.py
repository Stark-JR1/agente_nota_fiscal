from __future__ import annotations

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


ABA_PROCESSOS = "PROCESSOS_ENVIADOS"
ABA_RODOPAR = "RODOPAR_ABERTO"
ABA_CONFERENCIA = "CONFERENCIA"
ARQUIVO_CONTROLE = "CONTROLE_FISCAL.xlsx"
STATUS_FISCAL_INICIAL = "PRONTO_PARA_ENVIO"
STATUS_PROCESSOS_ENVIADOS = {"APROVADO", "APROVADO_PARCIAL"}
STATUS_FISCAIS = ["PRONTO_PARA_ENVIO", "ENVIADO_FISCAL", "EM_LANCAMENTO", "LANCADO", "CANCELADO"]

COLUNAS = [
    "Data Processamento",
    "Pedido",
    "Fornecedor",
    "CNPJ",
    "NF",
    "Valor",
    "Status Processo",
    "Arquivo Completo",
    "Status Fiscal",
    "Data Envio Fiscal",
    "Observação",
]

COLUNAS_RODOPAR = ["Pedido", "Fornecedor", "Valor", "Data", "Observação"]
COLUNAS_CONFERENCIA = [
    "Pedido",
    "Fornecedor Rodopar",
    "Valor Rodopar",
    "Status no Controle",
    "NF",
    "Arquivo Completo",
    "Situação",
]
COLUNAS_MANUAIS = {"Status Fiscal", "Data Envio Fiscal", "Observação", "ObservaÃ§Ã£o"}


def atualizar_controle_fiscal(processos: list[dict], pasta_relatorios_base, data_processamento=None) -> Path:
    caminho = caminho_controle_fiscal(pasta_relatorios_base)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = carregar_controle_existente(caminho)
        criar_ou_atualizar_planilha(wb, processos, data_processamento or datetime.now())
        wb.save(caminho)
    except Exception:
        logging.exception("Falha ao atualizar CONTROLE_FISCAL.xlsx.")
    return caminho


def caminho_controle_fiscal(pasta_relatorios_base) -> Path:
    base_operacional = getattr(pasta_relatorios_base, "base", None)
    if base_operacional is not None:
        return Path(base_operacional) / "RELATORIOS" / ARQUIVO_CONTROLE
    return Path(pasta_relatorios_base) / ARQUIVO_CONTROLE


def carregar_controle_existente(caminho: Path):
    return load_workbook(caminho) if caminho.exists() else Workbook()


def criar_ou_atualizar_planilha(wb, processos: list[dict], data_processamento) -> None:
    ws_processos = _obter_aba(wb, ABA_PROCESSOS)
    _garantir_cabecalho(ws_processos, COLUNAS)

    indices = _indices_colunas(ws_processos)
    existentes = _mapear_linhas_existentes(ws_processos, indices)

    for processo in processos:
        if processo.get("status") not in STATUS_PROCESSOS_ENVIADOS:
            continue

        linha = montar_linha_controle(processo, data_processamento)
        chave = chave_linha(linha)
        row_idx = existentes.get(chave)
        if row_idx:
            _atualizar_linha_existente(ws_processos, row_idx, indices, linha)
            continue

        ws_processos.append([linha.get(coluna) for coluna in COLUNAS])
        existentes[chave] = ws_processos.max_row

    ws_rodopar = _obter_aba(wb, ABA_RODOPAR)
    _garantir_cabecalho(ws_rodopar, COLUNAS_RODOPAR)
    _atualizar_conferencia(wb)

    _formatar_planilha(wb[ABA_PROCESSOS], COLUNAS)
    _formatar_planilha(wb[ABA_RODOPAR], COLUNAS_RODOPAR)
    _formatar_planilha(wb[ABA_CONFERENCIA], COLUNAS_CONFERENCIA)
    _aplicar_validacao_status(wb[ABA_PROCESSOS])


def montar_linha_controle(processo: dict, data_processamento) -> dict[str, Any]:
    docs = processo.get("documentos", [])
    pedido = _juntar_unicos(doc.get("numero_pedido") for doc in docs)
    fornecedor = _primeiro(doc.get("fornecedor_nome") for doc in docs)
    cnpj = _primeiro(doc.get("fornecedor_cnpj") for doc in docs)
    nf = _juntar_unicos(doc.get("numero_nf") for doc in docs)
    valor = processo.get("valor_nf") or _primeiro(doc.get("valor_total") for doc in docs)

    return {
        "Data Processamento": data_processamento,
        "Pedido": pedido,
        "Fornecedor": fornecedor,
        "CNPJ": cnpj,
        "NF": nf,
        "Valor": _normalizar_valor_saida(valor),
        "Status Processo": processo.get("status"),
        "Arquivo Completo": processo.get("arquivo_unido"),
        "Status Fiscal": STATUS_FISCAL_INICIAL,
        "Data Envio Fiscal": None,
        "Observação": None,
    }


def gerar_chave_controle(processo: dict) -> tuple[str, str, str, str, str]:
    return chave_linha(montar_linha_controle(processo, datetime.now()))


def chave_linha(linha: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        _normalizar_texto_chave(linha.get("Pedido")),
        _normalizar_texto_chave(linha.get("NF")),
        _normalizar_texto_chave(linha.get("Fornecedor")),
        _normalizar_valor_chave(linha.get("Valor")),
        _normalizar_texto_chave(linha.get("Arquivo Completo")),
    )


def _obter_aba(wb, nome: str):
    if nome in wb.sheetnames:
        return wb[nome]
    if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None:
        ws = wb.active
        ws.title = nome
        return ws
    return wb.create_sheet(nome)


def _garantir_cabecalho(ws, colunas: list[str]) -> None:
    cabecalho_atual = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    if not any(cabecalho_atual):
        for col, coluna in enumerate(colunas, start=1):
            ws.cell(row=1, column=col, value=coluna)
        return

    for coluna in colunas:
        if coluna not in cabecalho_atual:
            ws.cell(row=1, column=len(cabecalho_atual) + 1, value=coluna)
            cabecalho_atual.append(coluna)


def _indices_colunas(ws) -> dict[str, int]:
    indices = {
        str(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=col).value
    }
    if "ObservaÃ§Ã£o" in indices and "Observação" not in indices:
        indices["Observação"] = indices["ObservaÃ§Ã£o"]
    return indices


def _mapear_linhas_existentes(ws, indices: dict[str, int]) -> dict[tuple[str, str, str, str, str], int]:
    existentes = {}
    for row_idx in range(2, ws.max_row + 1):
        linha = {coluna: ws.cell(row=row_idx, column=indices[coluna]).value for coluna in COLUNAS if coluna in indices}
        chave = chave_linha(linha)
        if any(chave):
            existentes[chave] = row_idx
    return existentes


def _atualizar_linha_existente(ws, row_idx: int, indices: dict[str, int], linha: dict[str, Any]) -> None:
    for coluna, valor in linha.items():
        if coluna in COLUNAS_MANUAIS:
            continue
        ws.cell(row=row_idx, column=indices[coluna], value=valor)

    status_fiscal_col = indices["Status Fiscal"]
    if ws.cell(row=row_idx, column=status_fiscal_col).value in (None, ""):
        ws.cell(row=row_idx, column=status_fiscal_col, value=STATUS_FISCAL_INICIAL)


def _atualizar_conferencia(wb) -> None:
    ws_rodopar = _obter_aba(wb, ABA_RODOPAR)
    ws_processos = _obter_aba(wb, ABA_PROCESSOS)
    ws = _obter_aba(wb, ABA_CONFERENCIA)
    ws.delete_rows(1, ws.max_row)
    ws.append(COLUNAS_CONFERENCIA)

    idx_rodopar = _indices_colunas(ws_rodopar)
    idx_processos = _indices_colunas(ws_processos)
    processos_por_pedido: dict[str, list[int]] = {}
    for row_idx in range(2, ws_processos.max_row + 1):
        pedido = _valor(ws_processos, row_idx, idx_processos, "Pedido")
        if pedido in (None, ""):
            continue
        processos_por_pedido.setdefault(_normalizar_texto_chave(pedido), []).append(row_idx)

    for row_idx in range(2, ws_rodopar.max_row + 1):
        pedido = _valor(ws_rodopar, row_idx, idx_rodopar, "Pedido")
        fornecedor = _valor(ws_rodopar, row_idx, idx_rodopar, "Fornecedor")
        valor = _valor(ws_rodopar, row_idx, idx_rodopar, "Valor")
        encontrados = processos_por_pedido.get(_normalizar_texto_chave(pedido), [])
        if not encontrados:
            ws.append([pedido, fornecedor, valor, None, None, None, "NÃO ENCONTRADO NO CONTROLE"])
            continue
        for processo_row in encontrados:
            ws.append(
                [
                    pedido,
                    fornecedor,
                    valor,
                    _valor(ws_processos, processo_row, idx_processos, "Status Fiscal"),
                    _valor(ws_processos, processo_row, idx_processos, "NF"),
                    _valor(ws_processos, processo_row, idx_processos, "Arquivo Completo"),
                    "JÁ GERADO",
                ]
            )


def _valor(ws, row_idx: int, indices: dict[str, int], coluna: str):
    indice = indices.get(coluna)
    if not indice:
        return None
    return ws.cell(row=row_idx, column=indice).value


def _formatar_planilha(ws, colunas: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)

    indices = _indices_colunas(ws)
    for coluna, indice in indices.items():
        largura = len(str(coluna)) + 2
        for (cell,) in ws.iter_rows(min_col=indice, max_col=indice, min_row=2):
            valor = cell.value
            if valor not in (None, ""):
                largura = max(largura, min(len(str(valor)) + 2, 60))
        ws.column_dimensions[ws.cell(row=1, column=indice).column_letter].width = max(largura, 12)

    for coluna_data in ("Data Processamento", "Data Envio Fiscal", "Data"):
        indice = indices.get(coluna_data)
        if indice:
            for (cell,) in ws.iter_rows(min_col=indice, max_col=indice, min_row=2):
                cell.number_format = "dd/mm/yyyy hh:mm:ss" if coluna_data == "Data Processamento" else "dd/mm/yyyy"

    for coluna_valor in ("Valor", "Valor Rodopar"):
        indice = indices.get(coluna_valor)
        if indice:
            for (cell,) in ws.iter_rows(min_col=indice, max_col=indice, min_row=2):
                cell.number_format = '"R$" #,##0.00'


def _aplicar_validacao_status(ws) -> None:
    indices = _indices_colunas(ws)
    indice = indices.get("Status Fiscal")
    if not indice:
        return
    intervalo = f"{ws.cell(row=2, column=indice).coordinate}:{ws.cell(row=1048576, column=indice).coordinate}"
    ws.data_validations.dataValidation = []
    validacao = DataValidation(type="list", formula1=f'"{",".join(STATUS_FISCAIS)}"', allow_blank=True)
    ws.add_data_validation(validacao)
    validacao.add(intervalo)


def _juntar_unicos(valores) -> str | None:
    unicos = sorted({str(valor).strip() for valor in valores if valor not in (None, "")})
    return ", ".join(unicos) if unicos else None


def _primeiro(valores):
    for valor in valores:
        if valor not in (None, ""):
            return valor
    return None


def _normalizar_texto_chave(valor) -> str:
    return " ".join(str(valor or "").strip().upper().split())


def _normalizar_valor_saida(valor):
    if valor in (None, ""):
        return None
    try:
        return float(Decimal(str(valor)))
    except (InvalidOperation, ValueError):
        return valor


def _normalizar_valor_chave(valor) -> str:
    if valor in (None, ""):
        return ""
    try:
        return f"{Decimal(str(valor)).quantize(Decimal('0.01'))}"
    except (InvalidOperation, ValueError):
        texto = str(valor).replace("R$", "").strip()
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        try:
            return f"{Decimal(texto).quantize(Decimal('0.01'))}"
        except InvalidOperation:
            return _normalizar_texto_chave(valor)
