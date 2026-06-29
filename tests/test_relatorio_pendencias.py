from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app.relatorio import gerar_relatorios


def test_relatorio_pendencias_enriquecido(tmp_path):
    processos = [
        {
            "id": "PROCESSO_001",
            "status": "PENDENTE_PEDIDO",
            "erros": ["Pedido de compra nao identificado."],
            "valor_nf": Decimal("100.00"),
            "valor_boletos": Decimal("100.00"),
            "valor_pedidos": None,
            "confianca_processo": 72,
            "score_auditoria": 85,
            "risco_auditoria": "MEDIO_RISCO",
            "criterios_score": ["NF encontrada", "Pedido nao encontrado"],
            "documentos": [
                {
                    "tipo_documento": "NF_PRODUTO",
                    "fornecedor_nome": "FORNECEDOR TESTE",
                    "fornecedor_cnpj": "12.345.678/0001-95",
                    "numero_nf": "123",
                    "valor_total": Decimal("100.00"),
                    "arquivo_nome": "FORNECEDOR - NF 123.pdf",
                    "arquivo_final": str(tmp_path / "PENDENCIAS" / "NF.pdf"),
                    "origem_texto": "digital",
                    "observacoes": "NF localizada",
                },
                {
                    "tipo_documento": "BOLETO",
                    "valor_total": Decimal("100.00"),
                    "arquivo_nome": "FORNECEDOR - BOLETO.pdf",
                    "origem_texto": "ocr",
                },
            ],
        }
    ]

    _geral, pendencias, _txt = gerar_relatorios(processos, tmp_path, datetime(2026, 6, 18))

    wb = load_workbook(pendencias)
    ws = wb["Pendencias"]
    cabecalho = [cell.value for cell in ws[1]]
    linha = {cabecalho[i]: ws.cell(row=2, column=i + 1).value for i in range(len(cabecalho))}

    assert "Motivo Principal" in cabecalho
    assert "Ação Recomendada" in cabecalho
    assert "Documentos Encontrados" in cabecalho
    assert "Documentos Faltantes" in cabecalho
    assert "Arquivos Envolvidos" in cabecalho
    assert "Diferença NF x Boleto" in cabecalho
    assert "Score Auditoria" in cabecalho
    assert "Risco Auditoria" in cabecalho
    assert "Critérios Score" in cabecalho
    assert linha["Motivo Principal"] == "Pedido de compra nao identificado"
    assert "Anexar o pedido de compra" in linha["Ação Recomendada"]
    assert linha["Documentos Encontrados"] == "BOLETO; NF_PRODUTO"
    assert linha["Documentos Faltantes"] == "PEDIDO_COMPRA"
    assert "FORNECEDOR - NF 123.pdf" in linha["Arquivos Envolvidos"]
    assert linha["Score Auditoria"] == 85
    assert linha["Risco Auditoria"] == "MEDIO_RISCO"
    assert "Pedido nao encontrado" in linha["Critérios Score"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
