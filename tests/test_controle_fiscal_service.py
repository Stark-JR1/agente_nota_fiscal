from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app.services.controle_fiscal_service import (
    ABA_CONFERENCIA,
    ABA_PROCESSOS,
    ABA_RODOPAR,
    ARQUIVO_CONTROLE,
    STATUS_FISCAL_INICIAL,
    atualizar_controle_fiscal,
)


def processo(status="APROVADO", pedido="456", nf="123", arquivo=r"C:\saida\FORNECEDOR - NF 123.pdf"):
    return {
        "status": status,
        "valor_nf": Decimal("100.00"),
        "arquivo_unido": arquivo,
        "documentos": [
            {
                "fornecedor_nome": "FORNECEDOR TESTE",
                "fornecedor_cnpj": "12.345.678/0001-95",
                "numero_nf": nf,
                "numero_pedido": pedido,
                "valor_total": Decimal("100.00"),
            }
        ],
    }


def linhas(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


def test_cria_controle_fiscal_com_abas_e_processo_aprovado(tmp_path):
    caminho = atualizar_controle_fiscal([processo()], tmp_path, datetime(2026, 6, 18, 8, 0, 0))

    assert caminho == tmp_path / ARQUIVO_CONTROLE
    wb = load_workbook(caminho)
    assert ABA_PROCESSOS in wb.sheetnames
    assert ABA_RODOPAR in wb.sheetnames
    assert ABA_CONFERENCIA in wb.sheetnames

    ws = wb[ABA_PROCESSOS]
    assert len(linhas(ws)) == 1
    assert ws["A2"].value == datetime(2026, 6, 18, 8, 0, 0)
    assert ws["B2"].value == "456"
    assert ws["C2"].value == "FORNECEDOR TESTE"
    assert ws["E2"].value == "123"
    assert ws["F2"].value == 100
    assert ws["I2"].value == STATUS_FISCAL_INICIAL
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws["A1"].font.bold
    assert ws["F2"].number_format == '"R$" #,##0.00'


def test_adiciona_aprovado_parcial_e_nao_adiciona_pendente(tmp_path):
    processos = [
        processo(status="APROVADO_PARCIAL", nf="124", arquivo=r"C:\saida\parcial.pdf"),
        processo(status="PENDENTE_VALOR", nf="999", arquivo=r"C:\saida\pendente.pdf"),
    ]

    caminho = atualizar_controle_fiscal(processos, tmp_path, datetime(2026, 6, 18, 8, 0, 0))

    ws = load_workbook(caminho)[ABA_PROCESSOS]
    assert len(linhas(ws)) == 1
    assert ws["G2"].value == "APROVADO_PARCIAL"
    assert ws["E2"].value == "124"


def test_nao_duplica_e_preserva_campos_manuais(tmp_path):
    processos = [processo()]
    caminho = atualizar_controle_fiscal(processos, tmp_path, datetime(2026, 6, 18, 8, 0, 0))

    wb = load_workbook(caminho)
    ws = wb[ABA_PROCESSOS]
    ws["I2"] = "ENVIADO_FISCAL"
    ws["J2"] = datetime(2026, 6, 19)
    ws["K2"] = "Observacao manual"
    wb.save(caminho)

    atualizar_controle_fiscal(processos, tmp_path, datetime(2026, 6, 18, 9, 0, 0))

    wb = load_workbook(caminho)
    ws = wb[ABA_PROCESSOS]
    assert len(linhas(ws)) == 1
    assert ws["A2"].value == datetime(2026, 6, 18, 9, 0, 0)
    assert ws["I2"].value == "ENVIADO_FISCAL"
    assert ws["J2"].value == datetime(2026, 6, 19)
    assert ws["K2"].value == "Observacao manual"


def test_chave_considera_arquivo_completo(tmp_path):
    processos = [
        processo(arquivo=r"C:\saida\arquivo_1.pdf"),
        processo(arquivo=r"C:\saida\arquivo_2.pdf"),
    ]

    caminho = atualizar_controle_fiscal(processos, tmp_path, datetime(2026, 6, 18, 8, 0, 0))

    ws = load_workbook(caminho)[ABA_PROCESSOS]
    assert len(linhas(ws)) == 2


def test_conferencia_compara_rodopar_com_processos_enviados(tmp_path):
    caminho = atualizar_controle_fiscal([processo()], tmp_path, datetime(2026, 6, 18, 8, 0, 0))

    wb = load_workbook(caminho)
    ws_rodopar = wb[ABA_RODOPAR]
    ws_rodopar.append(["456", "FORNECEDOR TESTE", 100, datetime(2026, 6, 18), None])
    ws_rodopar.append(["789", "OUTRO FORNECEDOR", 50, datetime(2026, 6, 18), None])
    wb.save(caminho)

    atualizar_controle_fiscal([processo()], tmp_path, datetime(2026, 6, 18, 9, 0, 0))

    wb = load_workbook(caminho)
    ws = wb[ABA_CONFERENCIA]
    assert ws["A2"].value == "456"
    assert ws["G2"].value == "JÁ GERADO"
    assert ws["A3"].value == "789"
    assert ws["G3"].value == "NÃO ENCONTRADO NO CONTROLE"
