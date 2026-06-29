from openpyxl import load_workbook

from app.services.controle_fiscal_service import ABA_PROCESSOS, ARQUIVO_CONTROLE
from app.tools import importar_completos_controle as importer


def criar_pdf_completo(tmp_path, nome="FORNECEDOR TESTE - NF 123 - R$ 100,00.pdf"):
    pasta = tmp_path / "PROCESSOS" / "2026" / "06-JUNHO" / "18-06-2026" / "COMPLETOS"
    pasta.mkdir(parents=True)
    pdf = pasta / nome
    pdf.write_bytes(b"%PDF-1.4\n")
    return pdf


def test_dry_run_nao_grava_planilha_e_resume(monkeypatch, tmp_path):
    criar_pdf_completo(tmp_path)
    monkeypatch.setattr(importer, "extrair_pedido_pdf", lambda _pdf: "456")

    resumo = importer.importar_completos_controle(tmp_path, ano=2026, mes=6, dry_run=True)

    assert resumo.arquivos_encontrados == 1
    assert resumo.novos_registros == 1
    assert resumo.duplicados == 0
    assert resumo.sem_pedido == 0
    assert not (tmp_path / "RELATORIOS" / ARQUIVO_CONTROLE).exists()


def test_importa_historico_sem_duplicar_e_preserva_manuais(monkeypatch, tmp_path):
    criar_pdf_completo(tmp_path)
    monkeypatch.setattr(importer, "extrair_pedido_pdf", lambda _pdf: "456")

    resumo = importer.importar_completos_controle(tmp_path, ano=2026, mes=6, dry_run=False)
    assert resumo.novos_registros == 1

    caminho = tmp_path / "RELATORIOS" / ARQUIVO_CONTROLE
    wb = load_workbook(caminho)
    ws = wb[ABA_PROCESSOS]
    ws["I2"] = "ENVIADO"
    ws["J2"] = "20/06/2026"
    ws["K2"] = "Manual"
    wb.save(caminho)

    resumo = importer.importar_completos_controle(tmp_path, ano=2026, mes=6, dry_run=False)
    wb = load_workbook(caminho)
    ws = wb[ABA_PROCESSOS]

    assert resumo.duplicados == 1
    assert ws.max_row == 2
    assert ws["A2"].value == "18/06/2026"
    assert ws["B2"].value == "456"
    assert ws["C2"].value == "FORNECEDOR TESTE"
    assert ws["E2"].value == "123"
    assert ws["F2"].value == 100
    assert ws["G2"].value == "APROVADO_HISTORICO"
    assert ws["I2"].value == "ENVIADO"
    assert ws["J2"].value == "20/06/2026"
    assert ws["K2"].value == "Manual"


def test_contabiliza_campos_ausentes(monkeypatch, tmp_path):
    criar_pdf_completo(tmp_path, "FORNECEDOR SEM DADOS - PROCESSO COMPLETO.pdf")
    monkeypatch.setattr(importer, "extrair_pedido_pdf", lambda _pdf: None)

    resumo = importer.importar_completos_controle(tmp_path, ano=2026, dry_run=True)

    assert resumo.arquivos_encontrados == 1
    assert resumo.sem_pedido == 1
    assert resumo.sem_nf == 1
    assert resumo.sem_valor == 1
